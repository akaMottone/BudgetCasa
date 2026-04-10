import pandas as pd
import numpy as np
import os
import re
from openpyxl import load_workbook
from datetime import datetime
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.pipeline import make_pipeline

# --------------------------------------------------
# CONFIGURAZIONE
# --------------------------------------------------
file_nuovo_nome = "CASA ROSSA Ale Nuovi movimenti.xlsx"
nome_foglio_mov = "Scheda Movimenti"
nome_foglio_cat = "Categorie"

def pulisci_percorso(p):
    return p.strip().lstrip('&amp;amp;amp;amp;amp;').strip().strip("'").strip('"').strip()

print("--- ASSISTENTE CATEGORIZZATORE PRO (V19.8 - COD CONTO + KEYWORDS + POS + BRAND da LUOGO) ---")

# --------------------------------------------------
# DIZIONI GENERICHE POS
# --------------------------------------------------
GENERIC_COD_CONTO = [
    "pagamento pos",
    "pagamento tramite pos",
    "pagamento tramite pos carta",
    "pagamento carta",
    "electronic pos",
    "pagamento electronic"
]

# Parole che NON aiutano a identificare un brand
STOPWORDS_MERCHANT = {
    "treviglio", "milano", "roma", "italia",
    "spa", "srl", "snc", "sas", "s.p.a",
    "pagamento", "tramite", "pos", "paypal"
}

# Brand forti manuali: questi restano sempre
BRAND_PRIORITARI_MANUALI = [
    "ipercoop", "coop", "zalando", "amazon", "prenatal",
    "ikea", "nike", "adidas", "decathlon", "esselunga"
]

# --------------------------------------------------
# PARAMETRI brand auto da Luogo (storico)
# --------------------------------------------------
AUTO_BRAND_MIN_OCCORRENZE = 3      # brand preso da Luogo se compare almeno N volte nello storico
AUTO_BRAND_MIN_LEN = 3            # lunghezza minima stringa
AUTO_BRAND_MAX = 40               # massimo brand auto estratti
AUTO_BRAND_SOLO_SENZA_SPAZI = True  # True = prende solo valori Luogo "single token" (es: "Coop", "Trenord")
                                   # False = può prendere anche "Frutta e verdura" ecc.

def is_cod_conto_generico(cod):
    cod = (cod or "").lower()
    return any(g in cod for g in GENERIC_COD_CONTO)

def normalizza_testo_brand(s: str) -> str:
    s = (s or "").lower().strip()
    s = s.replace("*", " ")
    s = re.sub(r"\b(it\*[\w\d]+|\d{6,})\b", " ", s)  # elimina IT*xxxx e numeroni
    s = re.sub(r"\s+", " ", s).strip()
    return s

def estrai_brand_prioritario(text: str, brand_list):
    """
    Cerca la presenza di un brand prioritario dentro il testo.
    Ritorna il brand (stringa) se trovato, altrimenti None.
    """
    t = " " + normalizza_testo_brand(text) + " "
    for b in brand_list:
        bb = " " + b + " "
        if bb in t:
            return b
    return None

def build_brand_prioritari_from_luogo(df_storico):
    """
    Estrae possibili 'brand' dalla colonna Luogo dello storico (valori frequenti).
    """
    if 'Luogo' not in df_storico.columns:
        return []

    luogo = df_storico['Luogo'].fillna('').astype(str).str.strip()
    luogo = luogo[luogo != '']
    luogo = luogo[luogo.str.lower() != 'nan']

    if luogo.empty:
        return []

    vc = luogo.value_counts()

    brands = []
    for val, cnt in vc.items():
        v = normalizza_testo_brand(val)

        if len(v) < AUTO_BRAND_MIN_LEN:
            continue
        if any(sw == v for sw in STOPWORDS_MERCHANT):
            continue
        if AUTO_BRAND_SOLO_SENZA_SPAZI and (" " in v):
            continue
        if cnt < AUTO_BRAND_MIN_OCCORRENZE:
            continue

        brands.append(v)
        if len(brands) >= AUTO_BRAND_MAX:
            break

    # rimuove duplicati mantenendo ordine
    seen = set()
    out = []
    for b in brands:
        if b not in seen:
            seen.add(b)
            out.append(b)
    return out

def get_merchant_effettivo(row, brand_list):
    """
    Se Cod conto è generico (POS) usa Dettaglio, altrimenti Cod conto.
    Poi se trova un brand prioritario nel testo, ritorna SOLO quel brand.
    """
    cod = str(row.get('Cod conto', '')).strip()
    det = str(row.get('Dettaglio', '')).strip()

    base = det if (is_cod_conto_generico(cod) and det) else cod
    base = normalizza_testo_brand(base)

    # Se dentro il testo c'è un brand prioritario, uso quello come merchant effettivo
    b = estrai_brand_prioritario(base, brand_list)
    if b:
        return b

    return base

def parse_importo_to_float(x):
    """
    Converte importi scritti come stringhe in float in modo robusto:
    - gestisce "1234.56", "1.234,56", "1,234.56", "-89.12€", ecc.
    - ritorna NaN se non convertibile
    """
    if x is None:
        return np.nan

    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return np.nan

    # rimuove simboli e spazi
    s = s.replace("€", "").replace(" ", "").replace("\u00a0", "")
    s = s.replace("−", "-")  # alcuni export usano il meno “strano”

    # se ci sono sia ',' che '.', capisco quale è il separatore decimale:
    # quello che appare più a destra
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            # virgola = decimali, punti = migliaia
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            # punto = decimali, virgole = migliaia
            s = s.replace(",", "")
    else:
        # solo virgola -> la tratto come decimale (rimuovo eventuali punti migliaia)
        if "," in s:
            s = s.replace(".", "")
            s = s.replace(",", ".")
        # solo punto -> decimale (rimuovo eventuali virgole migliaia)
        else:
            s = s.replace(",", "")

    try:
        return float(s)
    except:
        return np.nan


def formatta_importo_excel(file_path, sheet_name, col_name="Importo", fmt="#,##0.00"):
    """
    Imposta il number_format della colonna Importo in Excel.
    In ambiente/locale IT, '#,##0.00' verrà visualizzato come '#.##0,00'
    """
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.save(file_path)
        wb.close()
        return

    ws = wb[sheet_name]

    # trova la colonna per intestazione
    header = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
    if col_name not in header:
        wb.save(file_path)
        wb.close()
        return

    col_idx = header[col_name]
    max_row = ws.max_row

    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        # formatta solo se è numero
        if isinstance(cell.value, (int, float)) and cell.value is not None:
            cell.number_format = fmt

    wb.save(file_path)
    wb.close()

# --------------------------------------------------
# CHECK FILE
# --------------------------------------------------
if not os.path.exists(file_nuovo_nome):
    print(f"\nERRORE: Non trovo '{file_nuovo_nome}'!")
    exit()

path_raw = input("Trascina qui il file con lo storico dei movimenti: ")
path_storico = pulisci_percorso(path_raw)

# --------------------------------------------------
# 1. CARICAMENTO DATI
# --------------------------------------------------
print("\n[1/3] Caricamento file e categorie...")
df_storico = pd.read_excel(path_storico, sheet_name=nome_foglio_mov, dtype=str)
df_nuovo = pd.read_excel(file_nuovo_nome, sheet_name=nome_foglio_mov, dtype=str)
df_cat = pd.read_excel(path_storico, sheet_name=nome_foglio_cat, dtype=str)

# Mappa descrittiva categorie
mappa_info_cat = {
    str(r['Code']): f"{r['Categoria']} > {r['Sottocategoria']} ({r['Tipo']})"
    for _, r in df_cat.iterrows()
}

# --------------------------------------------------
# 1b. BRAND PRIORITARI = manuali + auto da Luogo storico
# --------------------------------------------------
brand_auto = build_brand_prioritari_from_luogo(df_storico)

# Unione manuali + auto (dedup)
BRAND_PRIORITARI = []
seen = set()
for b in (BRAND_PRIORITARI_MANUALI + brand_auto):
    bb = normalizza_testo_brand(b)
    if bb and bb not in seen:
        seen.add(bb)
        BRAND_PRIORITARI.append(bb)

if brand_auto:
    print("\n[INFO] Brand auto-estratti da 'Luogo' (storico):")
    print("  - " + "\n  - ".join(brand_auto[:30]) + ("" if len(brand_auto) <= 30 else "\n  ..."))
else:
    print("\n[INFO] Nessun brand auto-estratto da 'Luogo' (storico).")

# --------------------------------------------------
# 2. KEYWORD RULES (da foglio Categorie)
# --------------------------------------------------
def build_keyword_rules(df_cat):
    rules = []
    if 'Keywords' not in df_cat.columns:
        return rules

    for _, r in df_cat.iterrows():
        code = str(r.get('Code', '')).strip()
        kws = str(r.get('Keywords', '')).lower().strip()

        if not code or not kws or kws == 'nan':
            continue

        try:
            priority = int(r.get('Priority', 70))
        except:
            priority = 70

        parts = [p.strip() for p in re.split(r"[;,]", kws) if len(p.strip()) > 2]
        if not parts:
            continue

        pattern = r"(?:^|[\W_])(" + "|".join(re.escape(p) for p in parts) + r")(?:$|[\W_])"
        rx = re.compile(pattern, flags=re.IGNORECASE)
        rules.append((rx, code, priority))

    rules.sort(key=lambda x: x[2], reverse=True)
    return rules

keyword_rules = build_keyword_rules(df_cat)

def match_keyword(text):
    if not text:
        return None, None, None
    for rx, code, pr in keyword_rules:
        m = rx.search(text)
        if m:
            return code, pr, m.group(1)
    return None, None, None

# --------------------------------------------------
# 3. FEATURE PER MODELLO (SOLO MERCHANT EFFETTIVO)
# --------------------------------------------------
def prepara_testo(df):
    return df.apply(lambda r: get_merchant_effettivo(r, BRAND_PRIORITARI), axis=1)

# --------------------------------------------------
# 4. ALLENAMENTO
# --------------------------------------------------
print("[2/3] Allenamento Intelligenza Artificiale...")
X_train = prepara_testo(df_storico)

def crea_modello():
    return make_pipeline(
        TfidfVectorizer(lowercase=True, token_pattern=r"(?u)\b\w+\b"),
        RandomForestClassifier(n_estimators=100, random_state=42, n_jobs=-1)
    )

model_code = crea_modello().fit(X_train, df_storico['Code'].fillna('NON_DEF'))
model_luogo = crea_modello().fit(X_train, df_storico['Luogo'].fillna(''))
model_persona = crea_modello().fit(X_train, df_storico['Persona'].fillna(''))

# --------------------------------------------------
# 5. ANALISI NUOVI MOVIMENTI
# --------------------------------------------------
print("[3/3] Analisi nuovi movimenti...")
X_nuovo = prepara_testo(df_nuovo)
probs_code = model_code.predict_proba(X_nuovo)
classes_code = model_code.classes_

stats = {"auto": 0, "manual": 0, "saltate": 0, "paypal": 0, "luoghi_auto": 0, "persone_auto": 0}
stop_manuale = False

p_max_series = np.max(probs_code, axis=1)
cond_ia_incerta = p_max_series < 0.50
cond_senza_codice = df_nuovo['Code'].isna() | (df_nuovo['Code'].astype(str).str.strip().str.lower() == 'nan')
cond_no_paypal = ~df_nuovo['Dettaglio'].fillna('').str.lower().str.contains("paga in 3 rate|paypal *paga", na=False)

totale_manuali = len(df_nuovo[cond_ia_incerta & cond_senza_codice & cond_no_paypal])
contatore_manuale = 0

for i, row in df_nuovo.iterrows():
    dettaglio = str(row.get('Dettaglio', '')).lower()
    merchant_eff = get_merchant_effettivo(row, BRAND_PRIORITARI)

    # PayPal (regola di esclusione)
    if "paga in 3 rate" in dettaglio or "paypal *paga" in dettaglio:
        stats["paypal"] += 1
        continue

    # già classificato
    if pd.notna(row.get('Code')) and str(row.get('Code')).strip().lower() != 'nan':
        continue

    # ---------- KEYWORD ----------
    # Se vuoi includere Dettaglio nelle keyword (consigliato), lo aggiungo qui:
    kw_text = merchant_eff + " " + normalizza_testo_brand(row.get('Dettaglio', ''))
    kw_code, kw_pr, kw_hit = match_keyword(kw_text)

    if kw_code and kw_pr >= 80:
        df_nuovo.at[i, 'Code'] = kw_code
        df_nuovo.at[i, 'Machine learning'] = f"KEYWORD: {kw_hit}"
        stats["auto"] += 1
        continue

    # ---------- LUOGO / PERSONA ----------
    for model, col, key in [(model_luogo, 'Luogo', "luoghi_auto"), (model_persona, 'Persona', "persone_auto")]:
        probs_lp = model.predict_proba(X_nuovo.iloc[[i]])
        if np.max(probs_lp) >= 0.55:
            df_nuovo.at[i, col] = model.classes_[np.argmax(probs_lp)]
            stats[key] += 1

    # ---------- ML ----------
    p_values = probs_code[i]
    p_max = np.max(p_values)
    cod_ia = str(classes_code[np.argmax(p_values)])

    # ---------- STORICO ----------
    voci_storiche = []
    merchant_norm = normalizza_testo_brand(merchant_eff)
    tokens = [t for t in merchant_norm.split() if len(t) > 2][:3]

    if tokens:
        filtro = pd.Series(True, index=df_storico.index)
        storico_cod = prepara_testo(df_storico)
        for t in tokens:
            filtro &= storico_cod.str.contains(re.escape(t), na=False)
        voci_storiche = df_storico[filtro]['Code'].value_counts().head(3).index.tolist()

    # ---------- OUTPUT ----------
    if p_max >= 0.50:
        df_nuovo.at[i, 'Code'] = cod_ia
        df_nuovo.at[i, 'Machine learning'] = f"AI: OK ({int(p_max*100)}%)"
        stats["auto"] += 1

    elif not stop_manuale:
        contatore_manuale += 1
        print("\n" + "=" * 70)
        print(f"DATA: {str(row.get('Data',''))[:10]} | IMPORTO: {row.get('Importo','')}€")
        print(f"VOCE: {row.get('Cod conto','')}")
        print(f"        ↳ {row.get('Dettaglio','')}")
        print(f"MERCHANT eff.: {merchant_eff}")
        print(f"{contatore_manuale} di {totale_manuali}")
        print("-" * 70)

        opzioni = []

        if kw_code and kw_code not in opzioni:
            opzioni.append(kw_code)
            print(f"[{len(opzioni)}] {kw_code} | KEYWORD '{kw_hit}' | {mappa_info_cat.get(kw_code,'N/D')}")

        for idx in np.argsort(p_values)[::-1][:3]:
            c = str(classes_code[idx])
            if c not in opzioni:
                opzioni.append(c)
                print(f"[{len(opzioni)}] {c} | {int(p_values[idx]*100)}% | {mappa_info_cat.get(c,'N/D')}")

        for c in voci_storiche:
            if c not in opzioni:
                opzioni.append(c)
                print(f"[{len(opzioni)}] {c} (Storico) | {mappa_info_cat.get(c,'N/D')}")

        scelta = input("\n" + "[Invio] Salta | [q] Esci | Numero/codice: ").strip()

        if scelta.lower() == 'q':
            stop_manuale = True
            stats["saltate"] += 1
        elif scelta == "":
            stats["saltate"] += 1
        elif scelta.isdigit() and 1 <= int(scelta) <= len(opzioni):
            df_nuovo.at[i, 'Code'] = opzioni[int(scelta)-1]
            df_nuovo.at[i, 'Machine learning'] = "Manuale (Suggerito)"
            stats["manual"] += 1
        else:
            df_nuovo.at[i, 'Code'] = scelta.upper()
            df_nuovo.at[i, 'Machine learning'] = "Manuale (Nuovo)"
            stats["manual"] += 1
    else:
        stats["saltate"] += 1

# --------------------------------------------------
# 6. SALVATAGGIO + FIX DATA
# --------------------------------------------------
print("\nSalvataggio dei dati in corso...")

# --- FIX DATA: rimuove l'orario e lascia solo la data ---
if 'Data' in df_nuovo.columns:
    df_nuovo['Data'] = pd.to_datetime(df_nuovo['Data'], errors='coerce').dt.date

# --- FIX IMPORTO: converte in numero vero (float) ---
if 'Importo' in df_nuovo.columns:
    df_nuovo['Importo'] = df_nuovo['Importo'].apply(parse_importo_to_float)

with pd.ExcelWriter(file_nuovo_nome, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
    df_nuovo.to_excel(writer, sheet_name=nome_foglio_mov, index=False)

# --- FORMAT EXCEL: mostra con virgola e 2 decimali (locale IT) ---
formatta_importo_excel(file_nuovo_nome, nome_foglio_mov, col_name="Importo", fmt="#,##0.00")

print("\n" + "="*30)
print("LAVORO COMPLETATO!")
print("="*30)
print(f"✅ IA: {stats['auto']} | 👤 Manuali: {stats['manual']} | ⏭ Saltati: {stats['saltate']}")
print(f"💳 PayPal ignorati: {stats['paypal']}")
print(f"📍 Luoghi: {stats['luoghi_auto']} | Persone: {stats['persone_auto']}")
print("="*30)